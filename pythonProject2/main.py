import PySimpleGUI as sg
import openpyxl
import time
import sys
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.fernet import Fernet
import base64
import os

def change_file_extension(file_path, new_extension):
    base_name = os.path.splitext(file_path)[0]
    new_file_path = f"{base_name}.{new_extension}"

    if os.path.exists(new_file_path):
        os.remove(new_file_path)

    os.rename(file_path, new_file_path)

def generate_key(password):
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=b"salt",
        iterations=100000
    )
    return Fernet(base64.urlsafe_b64encode(kdf.derive(password.encode())))

def encrypt_file(file_path, password):
    f = generate_key(password)

    with open(file_path, 'rb') as infile:
        original_data = infile.read()

    encrypted_data = f.encrypt(original_data)

    with open(file_path, 'wb') as outfile:
        outfile.write(encrypted_data)

def decrypt_file(file_path, password):
    f = generate_key(password)

    with open(file_path, 'rb') as infile:
        encrypted_data = infile.read()

    decrypted_data = f.decrypt(encrypted_data)

    with open(file_path, 'wb') as outfile:
        outfile.write(decrypted_data)

def main():
    # Определяем путь к временному каталогу, созданному PyInstaller
    if getattr(sys, 'frozen', False):
        # Мы в режиме onefile. Путь к временному каталогу хранится в sys._MEIPASS
        temp_folder = sys._MEIPASS
    else:
        # Мы в режиме разработки. Путь к временному каталогу совпадает с текущим каталогом
        temp_folder = os.path.abspath(os.path.dirname(__file__))

        # Загружаем изображение из временного каталога
    image_path = os.path.join(temp_folder, 'list.png')
    with open(image_path, 'rb') as f:
        image_data = f.read()

    layout = [
        [sg.Image(data=image_data, key='-IMAGE-')],
        [sg.Button('Далее', key='-NEXT-', size=(10, 1), button_color=('white', 'red'), font=('Helvetica', 12))]
    ]

    window = sg.Window('', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        elif event == '-NEXT-':
            window.close()
            time.sleep(1)
            main_window()

    window.close()
def validate_time_input(time):
    try:
        time_float = float(time)
        return 0.1 <= time_float <= 24
    except ValueError:
        return False
def main_window():
    layout = [
        [sg.Text('Введите наименование подразделения ОБИ')],
        [sg.InputText(key='-ORG_NAME-')],
        [sg.Text('(Пример: отделение  ЗИ от НСД | СЗРТ ВВО)', font='Courier 10 italic')],
        [sg.Button('Ввести')],
    ]

    window = sg.Window('Ввод названияи', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        if event == 'Ввести':
            organization_name = values['-ORG_NAME-']
            if organization_name:
                window.close()
                specialists_window()
            else:
                sg.popup_error("Пожалуйста, введите.")
    window.close()

def specialists_window():
    layout = [
        [sg.Text('Введите количество специалистов:')],
        [sg.Text('(Согласно штата)', font='Courier 10 italic')],
        [sg.InputText(key='-NUM_SPECIALISTS-')],
        [sg.Button('Ввести')]
    ]

    window = sg.Window('Ввод количества специалистов', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        if event == 'Ввести':
            num_specialists = values['-NUM_SPECIALISTS-']
            if num_specialists.isdigit() and int(num_specialists)<50:
                window.close()
                task_types_window()
            else:
                sg.popup_error("Пожалуйста, введите целое число.")
    window.close()

def task_types_window():
    layout = [
        [sg.Text('Введите количество возложенных ТИПОВЫХ ')],
        [sg.Text('задач на подразделение ОБИ:')],
        [sg.Text('(Определены положением о подразделении)', font='Courier 10 italic')],
        [sg.InputText(key='-NUM_TASK_TYPES-')],
        [sg.Button('Ввести')],
    ]

    window = sg.Window('Ввод количества типовых задач', layout)

    while True:
        event, values = window.read()
        global num_task_types
        if event == sg.WINDOW_CLOSED:
            break
        if event == 'Ввести':
            num_task_types = values['-NUM_TASK_TYPES-']
            if num_task_types.isdigit():
                window.close()
                typical_tasks_params_window(int(num_task_types))
            else:
                sg.popup_error("Пожалуйста, введите целое число.")
    window.close()

def typical_tasks_params_window(num_task_types):
    layout = [
        [sg.Text(f'Введите параметры для каждой типовой задачи (1 - {num_task_types}):'),
         sg.Text('\t\t\t  Время выполнения одним специалистом (в часах)'),
         sg.Text('\t\tЧастота возникновения (в течение года 1/365)')],
        *[[sg.Text(f'Задача {i + 1}:'), sg.InputText(key=f'-TASK_NAME_{i + 1}-'),
           sg.Text('Время выполнения:'), sg.InputText(key=f'-TASK_TIME_{i + 1}-'),
           sg.Text('Частота:'), sg.InputText(key=f'-TASK_FREQ_{i + 1}-')] for i in range(num_task_types)],
        [sg.Button('Ввести')]
    ]

    window = sg.Window('Ввод параметров типовых задач', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        if event == 'Ввести':
            task_names = [values[f'-TASK_NAME_{i+1}-'] for i in range(num_task_types)]
            task_times = [values[f'-TASK_TIME_{i+1}-'] for i in range(num_task_types)]
            task_freqs = [values[f'-TASK_FREQ_{i+1}-'] for i in range(num_task_types)]
            # Проверка на целые числа от 1 до 365
            if all(freq.isdigit() and 1 <= int(freq) <= 365 for freq in task_freqs):
                # Проверка на время выполнения в диапазоне от 0.1 до 24 часов
                if all(validate_time_input(time) for time in task_times):
                    # Вы можете сделать что-то с введенными данными, например, сохранить их в переменных или передать другой функции
                    window.close()
                    # Создаем новый файл Excel
                    global wb
                    wb = openpyxl.Workbook()
                    # Выбираем первый лист
                    sheet = wb.create_sheet('Типовые задачи')

                    # Записываем заголовки в первую строку
                    sheet.append(['Название типовой задачи', 'Время выполнения одним специалистом (в часах)',
                                  'Частота возникновения (в течение года 1/365)'])

                    # Записываем данные о типовых функциях в файл Excel
                    for i in range(num_task_types):
                        sheet.append([values[f'-TASK_NAME_{i + 1}-'], values[f'-TASK_TIME_{i + 1}-'],
                                      values[f'-TASK_FREQ_{i + 1}-']])


                    nontypical_tasks_window()
                    break
                else:
                    sg.popup('Пожалуйста, введите время выполнения задачи в диапазоне до 24 часов.')
            else:
                sg.popup('Пожалуйста, введите целые числа от 1 до 365 для частоты возникновения.')
    window.close()

def nontypical_tasks_window():
    layout = [
        [sg.Text('Введите количество возникающих НЕТИПОВЫХ')],
        [sg.Text('задач в подразделении ОБИ:')],
        [sg.InputText(key='-NUM_NON_TYPICAL_TASKS-')],
        [sg.Button('Ввести')],
    ]

    window = sg.Window('Ввод количества нетиповых задач', layout)

    while True:
        event, values = window.read()
        global num_nontypical_tasks
        if event == sg.WINDOW_CLOSED:
            break
        if event == 'Ввести':
            num_nontypical_tasks = values['-NUM_NON_TYPICAL_TASKS-']
            if num_nontypical_tasks.isdigit():
                window.close()
                nontypical_tasks_params_window(int(num_nontypical_tasks))
            else:
                sg.popup_error("Пожалуйста, введите целое число.")
    window.close()
def nontypical_tasks_params_window(num_nontypical_tasks):
    layout = [
        [sg.Text(f'Введите параметры для каждой нетиповой задачи (1 - {num_nontypical_tasks}):'),
         sg.Text('\t\t\t  Время выполнения одним специалистом (в часах)'),
         sg.Text('\t\tЧастота возникновения (в течение года 1/365)')],
        *[[sg.Text(f'Задача {i + 1}:'), sg.InputText(key=f'-NON_TYPICAL_TASK_NAME_{i + 1}-'),
           sg.Text('Время выполнения:'), sg.InputText(key=f'-NON_TYPICAL_TASK_TIME_{i + 1}-'),
           sg.Text('Частота:'), sg.InputText(key=f'-NON_TYPICAL_TASK_FREQ_{i + 1}-')] for i in range(num_nontypical_tasks)],
        [sg.Button('Ввести')]
    ]

    window = sg.Window('Ввод параметров нетиповых задач', layout)

    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        if event == 'Ввести':
            task_names_2 = [values[f'-NON_TYPICAL_TASK_NAME_{i+1}-'] for i in range(num_nontypical_tasks)]
            task_times_2 = [values[f'-NON_TYPICAL_TASK_TIME_{i+1}-'] for i in range(num_nontypical_tasks)]
            task_freqs_2 = [values[f'-NON_TYPICAL_TASK_FREQ_{i+1}-'] for i in range(num_nontypical_tasks)]
            # Проверка на целые числа от 1 до 365
            if all(freq.isdigit() and 1 <= int(freq) <= 365 for freq in task_freqs_2):
                # Проверка на время выполнения в диапазоне от 0.1 до 24 часов
                if all(validate_time_input(time) for time in task_times_2):

                    # Создаем новый лист для нетиповых функций
                    sheet = wb.create_sheet('Нетиповые задачи')

                    # Записываем заголовки в первую строку
                    sheet.append(['Название нетиповой задачи', 'Время выполнения одним специалистом (в часах)',
                                  'Частота возникновения (в течение года 1/365)'])

                    # Записываем данные о нетиповых функциях в файл Excel
                    for i in range(num_nontypical_tasks):
                        sheet.append(
                            [values[f'-NON_TYPICAL_TASK_NAME_{i + 1}-'], values[f'-NON_TYPICAL_TASK_TIME_{i + 1}-'],
                             values[f'-NON_TYPICAL_TASK_FREQ_{i + 1}-']])
                    # Сохраняем файл Excel
                    wb.save('Otchet.xlsx')
                    input_file = "Otchet.xlsx"
                    output_extension = "txt"
                    change_file_extension(input_file, output_extension)
                    input_file = "Otchet.txt"
                    password = "-JUB_em5E-V5yc0k2YlqP6"
                    encrypt_file(input_file, password)
                    download()
                    break
                else:
                    sg.popup('Пожалуйста, введите время выполнения задачи в диапазоне до 24 часов.')
            else:
                sg.popup('Пожалуйста, введите целые числа от 1 до 365 для частоты возникновения.')
    window.close()

def download():
    layout = [
        [sg.Text('Выполняется расчет...')],
        [sg.ProgressBar(1000, orientation='h', size=(40, 20), key='progressbar')],
        [sg.Cancel(), sg.Button('Close', visible=False, key='Close')],
        [sg.Text('Направьте файл "Otchet.txt" по ЗСПД: OK@KVVU.mil.zs', font=('Helvetica', 12), justification='center')]
    ]
    window = sg.Window('Расчет', layout, finalize=True)
    # loop that would normally do something useful
    for i in range(1000):
        event, values = window.read(timeout=10)
        if event == sg.WIN_CLOSED or event == 'Cancel':
            break
        progress_bar = window['progressbar']
        progress_bar.UpdateBar(i + 1)

    if event != sg.WIN_CLOSED and event != 'Cancel':
        if int(num_task_types) < 10:
            sg.popup("Ваше подразделение не соответствует\n(переизбыток личного состава)",
                     background_color='red', text_color='white', font=('Helvetica', 24))
        else:
            sg.popup("Ваше подразделение соответствует\n(оптимальное количество личного состава)",
                     background_color='green', text_color='white', font=('Helvetica', 24))
    # удаляем кнопку отмены и добавляем кнопку закрытия
    window['Cancel'].update(visible=False)
    window['Close'].update(visible=True)
main()
